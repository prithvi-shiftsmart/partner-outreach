#!/usr/bin/env python3
"""
Run OP->S1C test cases through the live concierge prompts and capture real LLM responses.

Builds prompts from the actual flow files (system prompts, tool schemas, templates),
feeds them through Claude CLI, and writes results to an Excel workbook for manual review.

Usage:
  python3 scripts/run_s1c_tests.py                    # run all 82 tests
  python3 scripts/run_s1c_tests.py --model sonnet      # use sonnet instead of haiku
  python3 scripts/run_s1c_tests.py --filter S1A-1      # only run tests matching filter
  python3 scripts/run_s1c_tests.py --concurrency 2     # lower concurrency
  python3 scripts/run_s1c_tests.py --out path.xlsx      # custom output path
"""

from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────

WORKSPACE = Path(__file__).resolve().parent.parent
FLOWS_DIR = WORKSPACE / "flows"
CLAUDE_CLI = "/Users/prithvi/.local/bin/claude"
DEFAULT_OUTPUT = Path("/Users/prithvi/Downloads/OP_to_S1C_Test_Cases_Live.xlsx")
DEFAULT_MODEL = "haiku"
TIMEOUT_SEC = 120

# Add scripts dir to path so we can import the test case generator
sys.path.insert(0, str(WORKSPACE / "scripts"))
from generate_s1c_test_cases import all_test_cases  # noqa: E402

# ── Excel Styles (matching generate_s1c_test_cases.py) ─────────────

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PARTNER_FILL = PatternFill("solid", fgColor="EAF6FF")
CONCIERGE_FILL = PatternFill("solid", fgColor="F4F4F4")
TOOL_FILL = PatternFill("solid", fgColor="FFF8E1")
LIVE_FILL = PatternFill("solid", fgColor="E8F5E9")  # light green for live responses
EXPECTED_FILL = PatternFill("solid", fgColor="F4F4F4")
GOOD_FONT = Font(color="27AE60")
OKAY_FONT = Font(color="F39C12")
BAD_FONT = Font(color="E74C3C")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
)
WRAP = Alignment(vertical="top", wrap_text=True)

COLUMNS = [
    ("Test ID", 10),
    ("Flow", 12),
    ("Category", 30),
    ("Scenario", 40),
    ("System Opening Message", 50),
    ("Partner Response", 35),
    ("Expected Tool Call(s)", 30),
    ("Simulated Tool Response", 40),
    ("Expected Response", 55),
    ("Live Concierge Response", 60),
    ("Manual Assessment", 12),
    ("Things to Fix", 25),
    ("Notes", 30),
]


# ── Prompt assembly ────────────────────────────────────────────────

def read_dir_md(directory: Path) -> str:
    """Concatenate all .md files in a directory, sorted alphabetically."""
    if not directory.exists():
        return ""
    parts = []
    for f in sorted(directory.glob("*.md")):
        parts.append(f.read_text().strip())
    return "\n\n---\n\n".join(parts)


def read_tool_dir(directory: Path) -> str:
    """Read tool schema + description from a tool directory."""
    parts = []
    schema_path = directory / "schema.json"
    desc_path = directory / "description.md"
    examples_path = directory / "examples.md"
    if schema_path.exists():
        parts.append(f"### Tool Schema: {directory.name}\n```json\n{schema_path.read_text().strip()}\n```")
    if desc_path.exists():
        parts.append(f"### Tool Description: {directory.name}\n{desc_path.read_text().strip()}")
    if examples_path.exists():
        parts.append(f"### Tool Examples: {directory.name}\n{examples_path.read_text().strip()}")
    return "\n\n".join(parts)


def read_all_tools(tools_dir: Path) -> str:
    """Read all tool subdirectories in a tools/ folder."""
    if not tools_dir.exists():
        return ""
    parts = []
    for item in sorted(tools_dir.iterdir()):
        if item.is_dir() and not item.name.startswith("_"):
            tool_text = read_tool_dir(item)
            if tool_text:
                parts.append(tool_text)
    return "\n\n".join(parts)


def load_flow_context(flow_name: str) -> dict:
    """Load all prompts, tools, and templates for a given flow."""
    if flow_name == "op_to_s1a":
        flow_dir = FLOWS_DIR / "op_to_s1a"
    elif flow_name == "s1a_to_s1c":
        flow_dir = FLOWS_DIR / "s1a_to_s1c"
    else:
        return {"system": "", "tools": "", "templates": ""}

    system_prompts = read_dir_md(flow_dir / "prompts")
    common_tools = read_all_tools(FLOWS_DIR / "_common" / "tools")
    flow_tools = read_all_tools(flow_dir / "tools")
    templates = read_dir_md(flow_dir / "templates")

    tools_combined = common_tools
    if flow_tools:
        tools_combined += "\n\n" + flow_tools

    return {
        "system": system_prompts,
        "tools": tools_combined,
        "templates": templates,
    }


# Cache flow contexts so we don't re-read files for every test
_flow_cache: dict[str, dict] = {}


def get_flow_context(flow_name: str) -> dict:
    if flow_name not in _flow_cache:
        _flow_cache[flow_name] = load_flow_context(flow_name)
    return _flow_cache[flow_name]


def determine_flow_name(test_case: dict) -> str | list[str] | None:
    """Determine which flow(s) to load based on test case metadata."""
    flow = test_case["flow"]
    test_id = test_case["id"]

    if flow == "OP→S1A":
        return "op_to_s1a"
    elif flow == "S1A→S1C":
        return "s1a_to_s1c"
    elif flow == "Regression":
        # REG-1.x to REG-3.x use OP->S1A prompts
        match = re.match(r"REG-(\d+)\.", test_id)
        if match:
            group = int(match.group(1))
            if group <= 3:
                return "op_to_s1a"
            elif group in (4, 5):
                return ["op_to_s1a", "s1a_to_s1c"]  # both flows
            elif group >= 6:
                return None  # metadata verification, skip LLM
    return "op_to_s1a"  # fallback


def is_metadata_test(test_case: dict) -> bool:
    """Check if this is a REG-6.x metadata verification test (no LLM call)."""
    return test_case["flow"] == "Regression" and test_case["id"].startswith("REG-6.")


def build_prompt(test_case: dict) -> str:
    """Assemble the full prompt for a test case."""
    flow_names = determine_flow_name(test_case)
    if flow_names is None:
        return ""

    # Handle single flow or multiple flows
    if isinstance(flow_names, list):
        system_parts = []
        tools_parts = []
        templates_parts = []
        for fn in flow_names:
            ctx = get_flow_context(fn)
            system_parts.append(ctx["system"])
            tools_parts.append(ctx["tools"])
            templates_parts.append(ctx["templates"])
        system_text = "\n\n===\n\n".join(system_parts)
        tools_text = "\n\n".join(dict.fromkeys(tools_parts))  # dedupe while preserving order
        templates_text = "\n\n===\n\n".join(templates_parts)
    else:
        ctx = get_flow_context(flow_names)
        system_text = ctx["system"]
        tools_text = ctx["tools"]
        templates_text = ctx["templates"]

    # Build the full prompt
    sections = []

    sections.append("# SYSTEM PROMPT\n\n" + system_text)

    if tools_text:
        sections.append("# AVAILABLE TOOLS\n\n" + tools_text)

    if templates_text:
        sections.append("# MESSAGE TEMPLATES\n\n" + templates_text)

    # Conversation context
    conversation = []
    conversation.append("# CONVERSATION CONTEXT\n")
    conversation.append(f"## System Opening Message (already sent to partner):\n{test_case['opening']}")
    conversation.append(f"## Partner Response:\n{test_case['partner']}")

    # Tool responses if applicable
    tool_resp = test_case.get("tool_response", "N/A")
    if tool_resp and tool_resp != "N/A":
        conversation.append(
            f"## Tool Call Results\n"
            f"The following tool calls were made and returned these results:\n\n"
            f"Tool calls: {test_case['tool_calls']}\n\n"
            f"Results:\n{tool_resp}"
        )

    sections.append("\n\n".join(conversation))

    # Final instruction
    sections.append(
        "# INSTRUCTION\n\n"
        "Generate the concierge's next SMS reply. Just the message text, nothing else. "
        "Follow the system prompt, templates, and tool results exactly. "
        "Do not include any prefixes like 'concierge:' or 'reply:'. "
        "Do not wrap in quotes or code fences. "
        "Plain text only, no emojis."
    )

    return "\n\n---\n\n".join(sections)


# ── Claude CLI execution ───────────────────────────────────────────

async def run_claude(prompt: str, model: str) -> tuple[str, float, str | None]:
    """Call Claude CLI, return (response, latency_seconds, error_or_none)."""
    t0 = time.monotonic()
    try:
        proc = await asyncio.create_subprocess_exec(
            CLAUDE_CLI, "-p", "--model", model,
            stdin=asyncio.subprocess.PIPE,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        stdout, stderr = await asyncio.wait_for(
            proc.communicate(input=prompt.encode()),
            timeout=TIMEOUT_SEC,
        )
        latency = time.monotonic() - t0
        raw = stdout.decode("utf-8", errors="replace")
        err = stderr.decode("utf-8", errors="replace")

        if proc.returncode != 0 and not raw.strip():
            return "", latency, f"CLI error (code={proc.returncode}): {err[:500]}"

        return clean_output(raw), latency, None

    except asyncio.TimeoutError:
        latency = time.monotonic() - t0
        try:
            proc.kill()
        except Exception:
            pass
        return "", latency, f"TIMEOUT after {TIMEOUT_SEC}s"
    except Exception as e:
        latency = time.monotonic() - t0
        return "", latency, str(e)


def clean_output(raw: str) -> str:
    """Strip quotes, backticks, code fences, and whitespace from CLI output."""
    s = raw.strip()

    # Remove surrounding quotes
    if len(s) >= 2 and s[0] in ('"', "'", "“") and s[-1] in ('"', "'", "”"):
        s = s[1:-1].strip()

    # Remove code fences
    if s.startswith("```"):
        lines = s.split("\n")
        # Remove first line (```lang)
        lines = lines[1:]
        # Remove last line if it's just ```
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        s = "\n".join(lines).strip()

    # Remove single backtick wrapping
    if len(s) >= 2 and s[0] == "`" and s[-1] == "`" and s.count("`") == 2:
        s = s[1:-1].strip()

    return s


# ── Test runner ────────────────────────────────────────────────────

async def run_test(test_case: dict, model: str, sem: asyncio.Semaphore) -> dict:
    """Run a single test case and return the result dict."""
    test_id = test_case["id"]

    # Metadata-only tests (REG-6.x)
    if is_metadata_test(test_case):
        return {
            **test_case,
            "live_response": "N/A -- metadata verification (no LLM call)",
            "latency": 0.0,
            "error": None,
        }

    prompt = build_prompt(test_case)
    if not prompt:
        return {
            **test_case,
            "live_response": "N/A -- could not build prompt",
            "latency": 0.0,
            "error": "Empty prompt",
        }

    async with sem:
        response, latency, error = await run_claude(prompt, model)

    if error:
        return {
            **test_case,
            "live_response": f"ERROR: {error}",
            "latency": latency,
            "error": error,
        }

    return {
        **test_case,
        "live_response": response,
        "latency": latency,
        "error": None,
    }


async def run_all_tests(cases: list[dict], model: str, concurrency: int) -> list[dict]:
    """Run all test cases with concurrency control."""
    sem = asyncio.Semaphore(concurrency)
    total = len(cases)
    completed = 0
    results = []
    t0 = time.monotonic()

    async def run_and_report(tc: dict) -> dict:
        nonlocal completed
        result = await run_test(tc, model, sem)
        completed += 1
        status = "ok" if not result.get("error") else "ERR"
        if is_metadata_test(tc):
            status = "skip"
        latency = result["latency"]
        mark = {"ok": "✓", "ERR": "✗", "skip": "—"}[status]
        print(f"  Running {tc['id']}... ({latency:.1f}s) {mark}  [{completed}/{total}]")
        return result

    # Launch all tasks
    tasks = [run_and_report(tc) for tc in cases]
    gathered = await asyncio.gather(*tasks)

    elapsed = time.monotonic() - t0
    return list(gathered), elapsed


# ── Excel export ───────────────────────────────────────────────────

def write_excel(results: list[dict], output_path: Path):
    """Write results to Excel with styling matching generate_s1c_test_cases.py."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Live Test Results"

    # Header row
    for col_idx, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["id"]).alignment = WRAP
        ws.cell(row=row_idx, column=2, value=r["flow"]).alignment = WRAP
        ws.cell(row=row_idx, column=3, value=r["category"]).alignment = WRAP
        ws.cell(row=row_idx, column=4, value=r["scenario"]).alignment = WRAP
        ws.cell(row=row_idx, column=5, value=r["opening"]).alignment = WRAP
        ws.cell(row=row_idx, column=6, value=r["partner"]).alignment = WRAP
        ws.cell(row=row_idx, column=7, value=r["tool_calls"]).alignment = WRAP
        ws.cell(row=row_idx, column=8, value=r["tool_response"]).alignment = WRAP
        ws.cell(row=row_idx, column=9, value=r["concierge"]).alignment = WRAP  # expected
        ws.cell(row=row_idx, column=10, value=r.get("live_response", "")).alignment = WRAP  # live
        ws.cell(row=row_idx, column=11, value="").alignment = WRAP  # manual assessment
        ws.cell(row=row_idx, column=12, value="").alignment = WRAP  # things to fix
        ws.cell(row=row_idx, column=13, value=r.get("notes", "")).alignment = WRAP

        # Borders
        for col_idx in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER

        # Tool columns fill
        ws.cell(row=row_idx, column=7).fill = TOOL_FILL
        ws.cell(row=row_idx, column=8).fill = TOOL_FILL

        # Expected response fill
        ws.cell(row=row_idx, column=9).fill = EXPECTED_FILL

        # Live response fill
        ws.cell(row=row_idx, column=10).fill = LIVE_FILL

        ws.row_dimensions[row_idx].height = 80

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ── Summary ────────────────────────────────────────────────────────

def print_summary(results: list[dict], elapsed: float):
    """Print a summary of the test run."""
    total = len(results)
    errors = sum(1 for r in results if r.get("error"))
    skipped = sum(1 for r in results if is_metadata_test(r))
    ran = total - skipped
    succeeded = ran - errors
    avg_latency = 0
    latencies = [r["latency"] for r in results if r["latency"] > 0 and not r.get("error")]
    if latencies:
        avg_latency = sum(latencies) / len(latencies)

    print()
    print("=" * 60)
    print(f"  Total tests:     {total}")
    print(f"  Ran (LLM call):  {ran}")
    print(f"  Skipped (meta):  {skipped}")
    print(f"  Succeeded:       {succeeded}")
    print(f"  Errors:          {errors}")
    print(f"  Avg latency:     {avg_latency:.1f}s")
    print(f"  Total time:      {elapsed:.1f}s")
    print("=" * 60)

    if errors:
        print("\n  Failed tests:")
        for r in results:
            if r.get("error"):
                print(f"    {r['id']}: {r['error'][:100]}")


# ── Main ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Run OP->S1C test cases through live concierge prompts"
    )
    parser.add_argument("--model", default=DEFAULT_MODEL,
                        help="LLM model: haiku (default), sonnet, opus")
    parser.add_argument("--filter", default=None,
                        help="Only run tests whose ID contains this substring")
    parser.add_argument("--concurrency", type=int, default=4,
                        help="Max concurrent Claude CLI calls (default: 4)")
    parser.add_argument("--out", default=None,
                        help=f"Output xlsx path (default: {DEFAULT_OUTPUT})")
    args = parser.parse_args()

    output_path = Path(args.out) if args.out else DEFAULT_OUTPUT

    # Load test cases
    cases = all_test_cases()

    # Apply filter
    if args.filter:
        cases = [c for c in cases if args.filter.lower() in c["id"].lower()]
        if not cases:
            print(f"No test cases match filter '{args.filter}'")
            sys.exit(1)

    print(f"Partner Outreach -- Live S1C Test Runner")
    print(f"  Model:       {args.model}")
    print(f"  Tests:       {len(cases)}")
    print(f"  Concurrency: {args.concurrency}")
    print(f"  Output:      {output_path}")
    print("-" * 60)

    # Pre-load flow contexts so they're cached
    get_flow_context("op_to_s1a")
    get_flow_context("s1a_to_s1c")
    print("  Flow prompts loaded.")
    print()

    # Run tests
    results, elapsed = asyncio.run(run_all_tests(cases, args.model, args.concurrency))

    # Write Excel
    write_excel(results, output_path)
    print(f"\n  Results saved to {output_path}")

    # Summary
    print_summary(results, elapsed)


if __name__ == "__main__":
    main()
