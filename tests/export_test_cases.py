#!/usr/bin/env python3
"""
Export DG concierge test cases from dg-test-cases.md to Excel.

Reads the markdown, parses each test case table, and writes a formatted
Excel workbook with color-coded category rows and a summary sheet.

Usage:
    python3 tests/export_test_cases.py
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# --- Color scheme by category ---
CATEGORY_FILLS = {
    "DG-Only Zone": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),      # light green
    "CKP-Only Zone": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),     # light blue
    "Both-Deal Zone": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),     # light yellow
    "Edge Case": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),          # light orange
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
BOLD_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def parse_markdown(md_path: Path) -> list[dict]:
    """Parse the markdown tables and return a list of test case dicts."""
    text = md_path.read_text()
    cases = []

    # Find all markdown table rows that start with | DG- or | CKP- or | BD- or | EC-
    # Each table has columns: ID | Category | Scenario | Partner Message | Expected Behavior | Score
    table_row_pattern = re.compile(
        r"^\|\s*([A-Z]+-\d+)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|",
        re.MULTILINE,
    )

    for match in table_row_pattern.finditer(text):
        test_id = match.group(1).strip()
        category = match.group(2).strip()
        scenario = match.group(3).strip()
        partner_message = match.group(4).strip()
        expected_behavior = match.group(5).strip()
        # Score is intentionally empty

        # Clean up markdown formatting
        for field_name in ("scenario", "partner_message", "expected_behavior"):
            val = locals()[field_name]
            # Remove markdown inline code backticks
            val = val.replace("`", "")
            # Remove markdown bold markers
            val = val.replace("**", "")
            # Convert markdown italic to plain
            val = re.sub(r"\*(.*?)\*", r"\1", val)
            locals()[field_name] = val

        cases.append(
            {
                "id": test_id,
                "category": category,
                "scenario": scenario,
                "partner_message": partner_message,
                "expected_behavior": expected_behavior,
            }
        )

    return cases


def build_workbook(cases: list[dict], output_path: Path) -> None:
    """Build the Excel workbook with test cases and summary sheets."""
    wb = Workbook()

    # ── Sheet 1: Test Cases ──
    ws = wb.active
    ws.title = "Test Cases"

    headers = ["ID", "Category", "Scenario", "Partner Message", "Expected Behavior", "Score", "Notes"]
    col_widths = [8, 18, 45, 40, 70, 10, 30]

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Set column widths
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, case in enumerate(cases, start=2):
        values = [
            case["id"],
            case["category"],
            case["scenario"],
            case["partner_message"],
            case["expected_behavior"],
            "",  # Score — empty
            "",  # Notes — empty
        ]
        category = case["category"]
        row_fill = CATEGORY_FILLS.get(category)

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

        # Bold the ID column
        ws.cell(row=row_idx, column=1).font = BOLD_FONT

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:G{len(cases) + 1}"

    # Set row heights for readability
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, len(cases) + 2):
        ws.row_dimensions[row_idx].height = 80

    # ── Sheet 2: Summary ──
    ws_summary = wb.create_sheet(title="Summary")

    # Count by category
    category_counts: dict[str, int] = {}
    for case in cases:
        cat = case["category"]
        category_counts[cat] = category_counts.get(cat, 0) + 1

    # Header
    summary_headers = ["Category", "Count", "Color"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 15

    # Category rows in display order
    display_order = ["DG-Only Zone", "CKP-Only Zone", "Both-Deal Zone", "Edge Case"]
    row_idx = 2
    for cat in display_order:
        count = category_counts.get(cat, 0)
        cell_cat = ws_summary.cell(row=row_idx, column=1, value=cat)
        cell_count = ws_summary.cell(row=row_idx, column=2, value=count)
        cell_color = ws_summary.cell(row=row_idx, column=3, value="")

        for cell in (cell_cat, cell_count, cell_color):
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

        # Color swatch
        if cat in CATEGORY_FILLS:
            cell_color.fill = CATEGORY_FILLS[cat]

        row_idx += 1

    # Total row
    total = sum(category_counts.values())
    cell_total_label = ws_summary.cell(row=row_idx, column=1, value="Total")
    cell_total_count = ws_summary.cell(row=row_idx, column=2, value=total)
    for cell in (cell_total_label, cell_total_count):
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")

    # Scoring legend
    row_idx += 2
    ws_summary.cell(row=row_idx, column=1, value="Scoring Legend").font = BOLD_FONT
    row_idx += 1
    for label, desc in [
        ("Pass", "Concierge response fully matches expected behavior"),
        ("Partial", "Response mostly correct but missing key elements or has minor issues"),
        ("Fail", "Response is wrong, uses wrong KB, leaks cross-company content, or violates guardrails"),
    ]:
        ws_summary.cell(row=row_idx, column=1, value=label).font = BOLD_FONT
        ws_summary.cell(row=row_idx, column=2, value=desc).font = BODY_FONT
        ws_summary.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
        row_idx += 1

    ws_summary.freeze_panes = "A2"

    # Save
    wb.save(output_path)
    print(f"Wrote {len(cases)} test cases to {output_path}")
    print(f"  Categories: {', '.join(f'{cat} ({count})' for cat, count in category_counts.items())}")
    print(f"  Sheets: 'Test Cases', 'Summary'")


def main():
    script_dir = Path(__file__).resolve().parent
    md_path = script_dir / "dg-test-cases.md"
    xlsx_path = script_dir / "dg-test-cases.xlsx"

    if not md_path.exists():
        raise FileNotFoundError(f"Markdown file not found: {md_path}")

    cases = parse_markdown(md_path)
    if not cases:
        raise ValueError("No test cases parsed from markdown. Check the table format.")

    build_workbook(cases, xlsx_path)


if __name__ == "__main__":
    main()
